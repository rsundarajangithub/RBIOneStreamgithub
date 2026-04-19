"""Generate mappings.xlsx — the single source of truth for OneStream → Dashboard mappings.

Run once to create the template, then edit in Excel. After editing, run:
    python generate_mappings.py
to produce data/mappings.json consumed by BL_DASH.html.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="2F5496")
seg_fill = PatternFill("solid", fgColor="D6E4F0")
comp_fill = PatternFill("solid", fgColor="FFF2CC")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = thin_border

def auto_width(ws, ncols, nrows):
    for col in range(1, ncols + 1):
        maxw = 0
        for row in range(1, nrows + 1):
            v = ws.cell(row=row, column=col).value
            if v:
                maxw = max(maxw, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(maxw + 3, 40)

# ═══════════════════════════════════════════════════════════════════
#  Sheet 1: Segments
# ═══════════════════════════════════════════════════════════════════
ws_seg = wb.active
ws_seg.title = "Segments"
seg_headers = ["segment_id", "brand_name", "entity", "ud3"]
ws_seg.append(seg_headers)
segments = [
    ("bk-usc",  "BK US&C",         "RestaurantBrandsIntlMNGT", "BK USC"),
    ("plk-usc", "Popeyes US&C",    "RestaurantBrandsIntlMNGT", "PLK USC"),
    ("fhs-usc", "Firehouse US&C",  "RestaurantBrandsIntlMNGT", "FHS USC"),
    ("th-usc",  "Tim Hortons C&US","RestaurantBrandsIntlMNGT", "TH C&US"),
    ("intl",    "RBI International","RestaurantBrandsIntlMNGT", "RBI_International"),
]
for row in segments:
    ws_seg.append(row)
style_header(ws_seg, len(seg_headers))
auto_width(ws_seg, len(seg_headers), len(segments) + 1)

# ═══════════════════════════════════════════════════════════════════
#  Sheet 2: Mappings  (the main config)
# ═══════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Mappings")
headers = [
    "segment",          # A — segment_id from Segments sheet
    "section",          # B — where it appears: kpi_card | kpi_tracker | financial | yoy | outlook | headline
    "dashboard_label",  # C — exact text shown on dashboard (must match JSON data)
    "os_account",       # D — OneStream account dimension
    "ud1",              # E — primary ud1 (for Actual / PY scenarios)
    "ud1_forecast",     # F — alternate ud1 for LE / Budget (leave blank if same)
    "scale",            # G — divisor: 1 = raw, 1000000 = ÷1M for $M display
    "value_type",       # H — currency | percent | count | rate
    "computed_from",    # I — for computed rows: semicolon-separated dashboard_labels to sum
    "notes",            # J — free text
]
ws.append(headers)

# ── BK US&C mappings (the most complete) ─────────────────────────

bk = "bk-usc"

# KPI Cards
kpi_rows = [
    (bk, "kpi_card", "SSS",   "ReportedTotalCompSalesSYSTEMPct", "LOB_FPA", "", 1, "percent", "", "Same-store sales %"),
    (bk, "kpi_card", "SST",   "ReportedTotalCompTrafficSYSTEMPct","LOB_FPA", "", 1, "percent", "", "Same-store traffic %"),
    (bk, "kpi_card", "NRG",   "SystemNRG",                       "LOB_FPA", "", 1, "count",   "", "Net restaurant growth"),
    (bk, "kpi_card", "AOI",   "",                                 "",        "", 1000000, "currency","Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A", "Computed sum of AOI sub-items"),
    (bk, "kpi_card", "Unit %","",                                 "",        "", 1, "percent", "DERIVED:NRG/StoreCount", "= NRG ÷ SystemStoreCount"),
    (bk, "kpi_card", "SWSG",  "",                                 "",        "", 1, "percent", "DERIVED:FzSalesGrowth",  "= TotalFzSales act/py − 1"),
    (bk, "kpi_card", "AOIG",  "",                                 "",        "", 1, "percent", "DERIVED:AoiGrowth",      "= AOI act/py − 1"),
]

# KPI Tracker rows
tracker_rows = [
    (bk, "kpi_tracker", "sss", "ReportedTotalCompSalesSYSTEMPct", "LOB_FPA", "", 1, "percent", "", ""),
    (bk, "kpi_tracker", "sst", "ReportedTotalCompTrafficSYSTEMPct","LOB_FPA", "", 1, "percent", "", ""),
    (bk, "kpi_tracker", "nrg", "SystemNRG",                       "LOB_FPA", "", 1, "count",   "", ""),
    (bk, "kpi_tracker", "aoi", "",                                 "",        "", 1000000, "currency","Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A", "Computed sum"),
]

# Financial P&L rows
fin_rows = [
    (bk, "financial", "FZ Sales",                "TotalFzSales",              "LOB_FPA",       "", 1000000, "currency", "", ""),
    (bk, "financial", "Royalty Rate %",           "RoyaltyRatePrc_Calc",      "Franchise_FPA", "", 1,       "rate",     "", "Decimal ratio"),
    (bk, "financial", "Royalties",                "RoyaltiesNet",             "Franchise_FPA", "", 1000000, "currency", "", ""),
    (bk, "financial", "FZ & Successor Fees",      "Franchise & Successor Fees","Franchise_FPA","", 1000000, "currency", "", ""),
    (bk, "financial", "Amortized Fees",           "",                          "",             "", 1000000, "currency", "", "NEEDS MAPPING"),
    (bk, "financial", "Unamortized Fees",         "Unamortized Fees_FPA",     "Franchise_FPA", "", 1000000, "currency", "", ""),
    (bk, "financial", "Total Franchise Revenue",  "FranchiseRevenueTotal",    "Franchise_FPA", "", 1000000, "currency", "", ""),
    (bk, "financial", "FZ Bad Debt",              "BadDebt",                  "Franchise_FPA", "", 1000000, "currency", "", ""),
    (bk, "financial", "Other FZ Expense",         "",                          "",             "", 1000000, "currency", "", "NEEDS MAPPING"),
    (bk, "financial", "Franchise AOI",            "AOI",                      "Franchise_FPA", "Franchise", 1000000, "currency", "", ""),
    (bk, "financial", "Property AOI",             "AOI",                      "Property_FPA",  "Property",  1000000, "currency", "", ""),
    (bk, "financial", "Ad Fund AOI",              "AOI",                      "Advertising",   "Advertising",1000000,"currency", "", ""),
    (bk, "financial", "Company AOI",              "AOI",                      "Company",       "Company",   1000000, "currency", "", ""),
    (bk, "financial", "Tech Fund AOI",            "AOI",                      "LOB_Regional_Tech_FPA","LOB_Regional_Tech",1000000,"currency","",""),
    (bk, "financial", "Segment G&A",              "AOI",                      "CorporateLOB_FPA","CorporateLOB",1000000,"currency","","May also be 601 in some scenarios"),
    (bk, "financial", "Other Corp",               "",                          "",             "", 1000000, "currency", "", "NEEDS MAPPING"),
    (bk, "financial", "Total AOI",                "",                          "",             "", 1000000, "currency", "Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A", "Computed sum"),
]

# YoY metrics
yoy_rows = [
    (bk, "yoy", "SSS",      "ReportedTotalCompSalesSYSTEMPct",  "LOB_FPA", "", 1, "percent", "", ""),
    (bk, "yoy", "SST",      "ReportedTotalCompTrafficSYSTEMPct","LOB_FPA", "", 1, "percent", "", ""),
    (bk, "yoy", "NRG",      "SystemNRG",                        "LOB_FPA", "", 1, "count",   "", ""),
    (bk, "yoy", "AOI ($M)", "",                                  "",        "", 1000000, "currency","Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A","Computed sum"),
    (bk, "yoy", "Unit %",   "",                                  "",        "", 1, "percent", "DERIVED:NRG/StoreCount", ""),
    (bk, "yoy", "SWSG",     "",                                  "",        "", 1, "percent", "DERIVED:FzSalesGrowth",  ""),
    (bk, "yoy", "AOIG",     "",                                  "",        "", 1, "percent", "DERIVED:AoiGrowth",      ""),
]

# Outlook & Headline (AOI-based)
other_rows = [
    (bk, "outlook",  "AOI", "", "", "", 1000000, "currency", "Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A", "Full-year outlook uses computed AOI"),
    (bk, "headline", "AOI", "", "", "", 1000000, "currency", "Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A", "Headline variance = AOI act − LE"),
]

all_bk = kpi_rows + tracker_rows + fin_rows + yoy_rows + other_rows

# ── Stub rows for other segments (user to fill in) ──────────────
other_segs = [
    ("plk-usc", "Popeyes"),
    ("fhs-usc", "Firehouse"),
    ("th-usc",  "Tim Hortons"),
    ("intl",    "International"),
]
other_rows_all = []
for seg_id, name in other_segs:
    other_rows_all += [
        (seg_id, "kpi_card",    "SSS", "ReportedTotalCompSalesSYSTEMPct", "LOB_FPA", "", 1, "percent", "", f"{name} — verify ud1"),
        (seg_id, "kpi_card",    "SST", "ReportedTotalCompTrafficSYSTEMPct","LOB_FPA","", 1, "percent", "", f"{name} — verify ud1"),
        (seg_id, "kpi_card",    "NRG", "SystemNRG",                       "LOB_FPA", "", 1, "count",   "", f"{name} — verify ud1"),
        (seg_id, "kpi_card",    "AOI", "",                                 "",        "", 1000000, "currency","Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A", f"{name} — verify components"),
        (seg_id, "kpi_tracker", "sss", "ReportedTotalCompSalesSYSTEMPct", "LOB_FPA", "", 1, "percent", "", ""),
        (seg_id, "kpi_tracker", "sst", "ReportedTotalCompTrafficSYSTEMPct","LOB_FPA","", 1, "percent", "", ""),
        (seg_id, "kpi_tracker", "nrg", "SystemNRG",                       "LOB_FPA", "", 1, "count",   "", ""),
        (seg_id, "kpi_tracker", "aoi", "",                                 "",        "", 1000000, "currency","Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A",""),
        (seg_id, "financial",   "Royalties",  "RoyaltiesNet",             "Franchise_FPA","",1000000,"currency","", f"{name} — verify ud1"),
        (seg_id, "financial",   "Total AOI",  "",                          "",        "", 1000000, "currency","Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A","Computed sum"),
        (seg_id, "yoy",        "AOI ($M)",   "",                          "",        "", 1000000, "currency","Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A",""),
        (seg_id, "outlook",    "AOI",        "",                          "",        "", 1000000, "currency","Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A",""),
        (seg_id, "headline",   "AOI",        "",                          "",        "", 1000000, "currency","Franchise AOI;Property AOI;Ad Fund AOI;Company AOI;Tech Fund AOI;Segment G&A",""),
    ]

# Write all rows
all_rows = all_bk + other_rows_all
for row in all_rows:
    ws.append(row)

# ── Format ──────────────────────────────────────────────────────
style_header(ws, len(headers))

# Colour-band by segment
prev_seg = None
for r in range(2, len(all_rows) + 2):
    seg_val = ws.cell(row=r, column=1).value
    if seg_val != prev_seg and prev_seg is not None:
        # Add segment separator colour
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = seg_fill
    prev_seg = seg_val
    # Highlight computed rows
    comp_val = ws.cell(row=r, column=9).value  # computed_from column
    if comp_val and not comp_val.startswith("DERIVED"):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = comp_fill
    # Highlight NEEDS MAPPING
    notes_val = ws.cell(row=r, column=10).value
    if notes_val and "NEEDS MAPPING" in str(notes_val):
        ws.cell(row=r, column=10).fill = PatternFill("solid", fgColor="FF9999")

auto_width(ws, len(headers), len(all_rows) + 1)

# Freeze top row
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_rows) + 1}"

# ═══════════════════════════════════════════════════════════════════
#  Sheet 3: Scenarios  (reference — which OS scenarios map to which)
# ═══════════════════════════════════════════════════════════════════
ws_sc = wb.create_sheet("Scenarios")
sc_headers = ["dashboard_scenario", "onestream_scenarios", "notes"]
ws_sc.append(sc_headers)
sc_rows = [
    ("actual",  "Actual",                                      "Current period actuals"),
    ("le",      "Forecast_02_10;Working;Forecast_01_11_MARKET","Latest Estimate — try in order"),
    ("budget",  "Budget",                                      "Annual operating plan"),
    ("py",      "Act_PY1_ActRate;Act_pY1_ActRate",            "Prior year actual at actual rate"),
]
for row in sc_rows:
    ws_sc.append(row)
style_header(ws_sc, len(sc_headers))
auto_width(ws_sc, len(sc_headers), len(sc_rows) + 1)

# ═══════════════════════════════════════════════════════════════════
#  Sheet 4: Time Periods  (reference)
# ═══════════════════════════════════════════════════════════════════
ws_tp = wb.create_sheet("Time Periods")
tp_headers = ["dashboard_view", "current_time", "py_time", "notes"]
ws_tp.append(tp_headers)
tp_rows = [
    ("march", "2026M3",  "2025M3",  "Monthly period (view=PER)"),
    ("q1",    "2026Q1",  "2025Q1",  "Quarter (view=PER)"),
    ("ytd",   "2026",    "2025",    "Year-to-date (view=PER)"),
]
for row in tp_rows:
    ws_tp.append(row)
style_header(ws_tp, len(tp_headers))
auto_width(ws_tp, len(tp_headers), len(tp_rows) + 1)

# ── Save ────────────────────────────────────────────────────────
outpath = "mappings.xlsx"
wb.save(outpath)
print(f"Created {outpath} with {len(all_rows)} mapping rows")
print("Sheets: Segments, Mappings, Scenarios, Time Periods")
print("\nNext: edit mappings.xlsx, then run: python generate_mappings.py")
